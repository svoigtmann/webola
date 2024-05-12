from PyQt5.Qt import QMessageBox, QIcon
import openpyxl
from pathlib import Path
from pony import orm
import re
import sys

from webola import database
from webola.database import db


def cell_from_sheet(sheet, row, col, key=None, allow_none=False):
    column = col[key] if key else col
    val = sheet.cell(row=row, column=column).value
    if val is None:
        print(f"[webola] Missing value in column '{key}' at row {row}")
        if allow_none:
            return '' # <UNKNOWN>
        else:
            sys.exit()
    else:
        return val.strip() if isinstance(val, str) else val

def create_dm23_db_lauf(nr, wettkampf, sheet, row, col, max_row):
    cell = lambda row, key: cell_from_sheet(sheet, row, col, key)
    startzeit = cell_from_sheet(sheet, row, col=2)    
    l    = database.Lauf(name      = "Lauf %d" % nr, 
                         wettkampf = wettkampf     ,
                         wettkampf_tag = 1,
                         tab_position = nr-1, 
                         startzeit = startzeit, finallauf=False)
    
    for n in range(1,max_row-row):
        first  = (sheet.cell(row=row+n, column=1).value or '').strip()
        second = (sheet.cell(row=row+n, column=2).value or '').strip()
        if re.match(r'\d\d-\d\d', first):
            nummer = cell(row+n, 'Startnummer')
            num    = int(nummer.split('-')[1])
            name   = cell(row+n, 'Vorname')+' '+cell(row+n, 'Nachname')
            verein = cell(row+n, 'Verein')
            klasse = cell(row+n, 'Klasse')

            wertung = verein not in ('Team Poland', 'Czech Team')
            if not wertung: 
                print(f"[webola] WARNING {name:<22} starts outside of the competition due to verein == '{verein}'." ) 
            
            t = database.Team(nummer=num, lauf=l, wertung=wertung)
            starter = database.Starter(name=name or "", verein=verein or "", klasse=klasse or "",team=t, nummer=1, strafen=0)
            starter.einheit = starter.get_einheit()
            
        elif second.startswith('Staffel'):
            l.titel = second

    db.commit()

def create_db_lauf(nr, wettkampf, sheet, row, col, max_row):
    cell = lambda row, key: cell_from_sheet(sheet, row, col, key, allow_none=True)    
    l    = database.Lauf(name      = "Lauf %d" % nr, 
                         wettkampf = wettkampf     ,
                         wettkampf_tag = 1,
                         tab_position = nr-1, 
                         startzeit = str(cell(row,'Startzeit')), finallauf=False)
    
    for n in range(max_row-row):
        name   = cell(row+n, 'Name'  )
        verein = cell(row+n, 'Verein')
        klasse = cell(row+n, 'Klasse')
        if all([name,verein,klasse]): 
            t = database.Team(nummer=n+1,lauf=l,wertung=True)
            starter = database.Starter(name=name or "", verein=verein or "", klasse=klasse or "",team=t, nummer=1, strafen=0)
            starter.einheit = starter.get_einheit()
        elif any([name,verein,klasse]): 
            info(f"*** Unvollständige Daten *** name='{name}', verein='{verein}', klasse='{klasse}' in Zeile {row+n}")

    db.commit()

def create_coloured_db_lauf(wettkampf, sheet, run_num, start, stop, row, col_for):
    cell = lambda row, key: cell_from_sheet(sheet, row, col_for, key, allow_none=True)
    l = database.Lauf(name      = "Lauf %d" % run_num, 
                      wettkampf = wettkampf    ,
                      wettkampf_tag = 1, 
                      tab_position = run_num,
                      startzeit = str(cell(row,'Time')), finallauf=False)
    
    for row in range(start,stop+1):
        nummer    = cell(row, 'Nr'        )
        firstname = cell(row, 'First name')
        lastname  = cell(row, 'Surname'   )
        if firstname and lastname:
            name      = " ".join([firstname, lastname])
            verein    = cell(row, 'Club')
            gender    = cell(row, 'Gender').upper()
            ageclass  = cell(row, 'Age classes')
            bow       = cell(row, 'Bow')
            klasse    = f"{ageclass} ({gender}) {bow}"

            t = database.Team(nummer=nummer,lauf=l,wertung=True)
            starter = database.Starter(name=name, verein=verein, klasse=klasse, team=t, nummer=1, strafen=0)
            starter.einheit = starter.get_einheit()

    db.commit()

    
def strip(label):
    text = label.strip()
    if text.startswith('Startliste'):
        text = text[10:].lstrip(': ')
    return text if len(text) else None

def xlsx2sql(filename, column=0):
    info(f"Reading '{filename}'")
    wb = openpyxl.load_workbook(filename, data_only=True)#, read_only=True)
    first = wb.sheetnames[0]
    sheet = wb[first]

    name = sheet.cell(row=1, column=1).value
    name = strip(name) if isinstance(name, str) else None
    wettkampf = database.Wettkampf.create(name)
    
    info(f"Found '{wettkampf.name}'")
    
    cols = list(sheet.columns)
    row_for, runs = parse_xlsx_column(cols[column])
        
    if is_dm_23_format(sheet, runs):
        return process_dm_23_startlist(sheet, row_for, runs, wettkampf)
    elif 'Lauf' in row_for.keys():
        return process_standard_startlist(sheet, row_for, runs, wettkampf)
    else:
        row_for, runs = parse_xlsx_column(cols[1])
        if 'Run' in row_for.keys():
            # note: runs = [ ... (run,start,invalid) ... ]
            runs = [ (run[0],run[1]) for run in runs ]
            return process_coloured_startlist(sheet, row_for, runs, wettkampf)
    
    error(f"I cannot determine format of '{filename.name}'. Please consider using 'startliste_dummy.xlsx' as a template.")

def is_dm_23_format(sheet, runs):
    if runs and len(runs[0]) > 1:
        run_id, row = runs[0][0:2]
        label = str(sheet.cell(row=row, column=1).value).strip()
        date  = str(sheet.cell(row=row, column=2).value).strip()
        return label == f"Lauf {run_id}" and 'Uhr' in date
    else:
        return False
    
def info(msg):
    print(f"[webola] {msg} ...")
        
def error(msg):    
    box = QMessageBox(QMessageBox.Critical, 'Webola', msg)
    box.setWindowIcon(QIcon(":/webola.png"))
    box.exec()
    sys.exit(1)
    
def process_coloured_startlist(sheet, row_for, runs, wettkampf):
    max_row = max(row_for.values())
    col_for = find_header_data(sheet, row_for['Run'], 'Run')
        
    for run_num, row in runs:
        start = find_rows_for_run(sheet, row, max_row, -1)
        stop  = find_rows_for_run(sheet, row, max_row, +1)
        create_coloured_db_lauf(wettkampf, sheet, run_num, start, stop, row, col_for)
    
    return wettkampf

def find_rows_for_run(sheet, row, max_row, offset):
    colour = lambda row: sheet.cell(row=row, column=2).fill.start_color.index
    start = colour(row)
    while colour(row) == start and row >=1 and row <= max_row:
        row = row + offset
    
    return row-offset

def process_dm_23_startlist(sheet, row_for, runs, wettkampf):
    # dm23 format has no headers ... fake it
    col_for = { v: k for k, v in enumerate(['Startnummer', 'Nachname', 'Vorname', 'Verein', 'Klasse'], 1) } 
    
    for run in runs:
        idx, start, stop = run[0], run[1], run[-1]
        create_dm23_db_lauf(idx, wettkampf, sheet, start, col_for, stop)

    info(f"Imported {len(runs)} runs")
    return wettkampf
    
def process_standard_startlist(sheet, row_for, runs, wettkampf):
    col_for = find_header_data(sheet, row_for['Lauf'], 'Lauf')
    for run,start,stop in runs:
        create_db_lauf(run, wettkampf, sheet, start, col_for, stop)

    return wettkampf

def entries2keys(it):
    simplify = lambda s: s if s == 'Lauf' else s.replace('Lauf','')
    value    = lambda c: maybe_int(simplify(c.value.strip())) if isinstance(c.value, str) else c.value
    return [ (val,row) for row, val in enumerate(map(value,it),1) if val is not None ]

def maybe_int(val):
    try:
        return int(val)
    except ValueError:
        return val 
        
def parse_xlsx_column(it): 
    row_for = {}
    runs    = []
    for key, row in entries2keys(it):
        if len(runs)>0: runs[-1].append(row-1)
        if isinstance(key,int):
            runs.append([key, row])
        else:
            row_for[key] = row
                
    if len(runs)>0 and len(runs[-1]) == 2:
        # maybe add max. row number to final entry 
        runs[-1].append(len(it))
            
    return row_for, runs

def find_header_data(sheet, row, key):
        if key == 'Run':
            need = "AN, Run, Time, Nr, Surname, First name, Nation, Club, Gender, Birthday, Age, Age classes, Bow" 
        else:
            need = "Lauf, Startzeit, Startnr., Name, Verein, Klasse"
            
        have = ", ".join([ c.value.strip() for c in sheet[row] if c.value is not None])
        ok   = need == have
                
        if not ok: 
            print('Need  headers: %s'        % need)
            print('Found headers: %s ... %s' % (have, "OK" if ok else "NOT OK" ))
            raise Exception('Header mismatch.')
              
        
        return { k: v for k,v in entries2keys(sheet[row]) } 


if __name__ == '__main__':
    
    filename = Path('../meldungen.xlsx').resolve()
    
    db.bind(provider='sqlite', filename=str(filename.with_suffix('.sql')), create_db=True)
    db.generate_mapping(create_tables=True)
        
    with orm.db_session:
    
        xlsx2sql(filename, column=0)


