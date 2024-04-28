import openpyxl
from openpyxl.styles import Font, Alignment, Border, Side
from openpyxl.styles.borders import BORDER_THIN
from openpyxl.styles.colors import BLACK
import time
from webola import utils
from webola.statistik import Medaillenspiegel, collect_data
from webola.exporter import Sheet, medaillenspiegel, generic_export
from webola.database import Team
from webola.utils import time2str
from PyQt5.Qt import QMessageBox, QIcon

STYLE = { 'arial'   : Font(name='Arial'),
          'huge'    : Font(name='Arial', size=22),
          'bold'    : Font(name='Arial', bold=True),
          'tiny'    : Font(name='Arial', size= 8),
          'center'  : Alignment(horizontal='center', vertical='center'),
          'vcenter' : Alignment(vertical  ='center'),
          'left'    : Alignment(vertical  ='center',horizontal='left'),
          'right'   : Alignment(vertical  ='center',horizontal='right') }

def get_sheet(header, wb):
    sheet = wb.create_sheet(header.replace('/',' ').replace(':',' '), 0)
    sheet.row_dimensions[1].height = 26
    write_cell = lambda r, c, t, *args: write_cell_to_sheet(sheet, r, c, t, STYLE['arial'], *args)
    return sheet, write_cell

def xls_export_zielliste(wettkampf, filename, header, tabs):
    try:
        wb = openpyxl.load_workbook(filename)
        for name in wb.sheetnames:
            sheet = wb[name]
            wb.remove(sheet)
    except:
        wb = openpyxl.Workbook()    
    
    ms = Medaillenspiegel(wettkampf)
    
    xlsx_export_serienbrief(wb, wettkampf, header, staffel=True )
    xlsx_export_serienbrief(wb, wettkampf, header, staffel=False)
    xlsx_export_runs       (wb, tabs)
    xlsx_export_medaillen  (wb, ms  )

    xlsx_export(wb, wettkampf, header, 'Ergebnis')

    try:
        wb.save(filename)
    except Exception as e:
        box = QMessageBox(QMessageBox.Critical, 'Export-Fehler',
                    f'Der Excel-Export war nicht möglich:<br><dd>{str(e)}</dd><br>' +
                    'Unter Windows kann z.B. keine bereits geöffnete Excel-Datei neu geschrieben werden. '+
                    'In diesem Fall muss Excel for dem Export geschlossen werden.<br>'+
                    'Es wird versucht, mögliche weitere Exporte dennoch zu erzeugen.')
        box.setWindowIcon(QIcon(":/webola.png"))
        box.exec()
        
    return ms

def xlsx_export_medaillen(wb, ms):
    sheet, write_cell = get_sheet('Medaillen', wb)
    
    toprule = lambda row, start=1, stop=5: create_toprule(sheet, row, start, stop)
    stand   = lambda row, start=1, stop=5: write_stand   (sheet, row, start, stop)
    
    width   = medaillenspiegel(ms, write_cell, toprule, stand, STYLE)

    sheet.column_dimensions[ 'B' ].width = width    

def xlsx_export_runs(wb, tabs):
    for idx,tab in reversed(list(tabs.enumerate_run_tabs())):
        name  = tabs.tab_name(idx)
        if name.endswith("'1") or name.endswith("'2"):
            name = name[:-2] + f' (Tag {name[-1]})'   
        
        head  = utils.join_nonempty(': ', name, tab.startinfo())
        if tab.is_running():
            name = name.replace('*','')
            head += " [AKTUELLES RENNEN]"
        xlsx_export(wb, tab.lauf, head, name)

def create_toprule(sheet, row, start=1,stop=9):
    border = Border(top=Side(border_style=BORDER_THIN, color=BLACK))
    for col in range(start,stop+1):
        sheet.cell(row=row,column=col).border = border

def write_stand(sheet, row, start=1, stop=9, align='right'):
    create_toprule(sheet, row+1,start=1, stop=stop)
    write_cell_to_sheet(sheet, row+1, start if align == 'left' else stop, 
                        'Stand: %s' % time.strftime("%d. %B %Y, %H:%M Uhr", time.localtime()),
                        STYLE['tiny'], STYLE[align])

def xlsx_export_serienbrief(wb, data, head, staffel):
    name = 'Serienbrief (%s)' % ("Staffel" if staffel else "Einzel")
    head = '%s: %s' % (name, head)
    sheet, write_cell = get_sheet(name, wb)
    serial_export(staffel, sheet, data, head, write_cell)
    
        
def xlsx_export(wb, data, head, name=None):

    sheet, write_cell = get_sheet(name, wb)
    
    toprule = lambda row, start=1, stop=9: create_toprule(sheet, row, start, stop)
    stand   = lambda row, start=1, stop=9: write_stand   (sheet, row, start, stop)
    
    generic_export(data, head, write_cell, toprule, stand, STYLE)
        
    for col,w in zip('ABCDEFGHI',(22,5,30,30,10,10,10,10,11)):
        sheet.column_dimensions[col].width = w
    
    sheet.sheet_view.showGridLines = False
    
def write_cell_to_sheet(sheet, r,c,text,*args):
    cell = sheet.cell(row=r, column=c)
    cell.value = text
    for s in args:
        if   isinstance(s, Font     ): cell.font      = s
        elif isinstance(s, Alignment): cell.alignment = s
        elif isinstance(s, Border   ): cell.border    = s
        else:
            raise ValueError("Unsopported style %s" % type(s))
        
def serial_export(staffel, xlsx, wettkampf, header, write_cell):
    
    n     = 1 if not staffel else max( t.anzahl() for l in wettkampf.laeufe for t in l.teams )
    sheet = Sheet(xlsx, header, write_cell, n)

    create_toprule(sheet.xlsx, 3, 1, sheet.max_col())
    create_toprule(sheet.xlsx, 4, 1, sheet.max_col())
    
    for wertung in collect_data(wettkampf):
        # renumber all positions within the current class since team.platz is the position in the team's run
        pos, sieger = 1, None
        
        for team in Team.sortiere(wertung.teams):
            if team.ist_staffel() == staffel and team.platz and not team.is_dsq():
                nr = team.anzahl()
                ns = team.lauf.anzahl_schiessen
                np = team.lauf.anzahl_pfeile
                fehler  = team.fehler() or 0
                treffer = nr*ns*np - fehler
                name, verein = team.get_name_verein()
                sheet.write('Klasse'     , wertung.klasse, newline=True)
                sheet.write('Name'       , name                  )
                sheet.write('Verein'     , verein                )
                sheet.write('Platz'      , "%d" % pos if team.wertung else "-" )
                sheet.write('Zeit'       , time2str(team.zeit()) )
                sheet.write('Fehler'     , fehler                )
                sheet.write('Treffer'    , treffer               )
                sheet.write('Laufzeit'   , time2str(team.laufzeit()))
                sheet.write('Strafen', sum(s.strafen for s in team.liste() ) )
                einheiten = set( s.einheit for s in team.liste() )
                if len(einheiten) == 1:
                    sheet.write('Einheit', einheiten.pop()   )
                else:
                    sheet.write('Einheit', 'div' )

                if pos == 1:
                    sieger = team.zeit()
                else:
                    sheet.write('Abstand', time2str(team.zeit() - sieger))

                if team.ist_staffel():
                    for idx,s in enumerate(team.liste(),1): 
                        fehler  = s.fehler  or 0
                        treffer = ns*np-fehler
                        sheet.write('Name'       , s.get_name()        , starter=idx)
                        sheet.write('Verein'     , s.verein            , starter=idx)
                        sheet.write('Klasse'     , s.klasse            , starter=idx)
                        sheet.write('Zeit'       , time2str(s.zeit())  , starter=idx)
                        sheet.write('Fehler'     , fehler              , starter=idx)
                        sheet.write('Treffer'    , treffer             , starter=idx)
                        sheet.write('Laufzeit'   , time2str(s.laufzeit), starter=idx)
                        sheet.write('Strafen'    , s.strafen           , starter=idx)
                        sheet.write('Einheit'    , s.einheit           , starter=idx)
                
                pos += 1 if team.wertung else 0

    write_stand(sheet.xlsx, sheet.row+1, 1, sheet.max_col(), align='left')
#    stand(row, stop=9)
        
        
