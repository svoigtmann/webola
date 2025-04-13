from webola.utils import time2str
from collections import defaultdict
from openpyxl.styles import Alignment, Font
from openpyxl.utils.cell import get_column_letter
from webola.statistik import collect_data
from webola.database import Team
from webola.latex import TexTableWriter, StaffelMode

class MockWriter():
    def __init__(self, write_cell):
        self.cell = lambda r, c, t, *args: write_cell(r, c, t, *args)
    
class Column():
    id         = 0
    write_cell = None
    xlsx       = None
    
    def __init__(self, name, width, *styles):
        Column.id  += 1
        self.num    = Column.id
        self.name   = name
        self.width  = width
        self.styles = styles
        
        Column.xlsx.column_dimensions[ get_column_letter(self.num) ].width = width 
        self.write_header(row=3)
    
    def copy(self, name):
        return Column(name, self.width+5, *self.styles)

    def write_header(self, row):
        Column.write_cell(row, self.num, self.name, Font(name='Arial', bold=True), *self.styles)
        return self

    def write(self, row, text):
        Column.write_cell(row, self.num, text, *self.styles)
        return self
        
class Sheet():
    def __init__(self, xlsx, header, writer, team):
        Column.id         = 0
        Column.write_cell = writer.cell
        Column.xlsx       = xlsx
        
        self.column    = dict()
        self.xlsx      = xlsx
        self.row       = 3
        self.n_staffel = 0
    
        writer.cell(1, 1, header, Font(name='Arial', size=22))
    
        self.add('Klasse'     , 20, 'left'  )
        self.add('Platz'      , 10, 'center')
        self.add('Name'       , 30, 'left'  )
        self.add('Verein'     , 30, 'left'  )
        self.add('Zeit'       , 10, 'center')
        self.add('Abstand'    , 10, 'center')
        self.add('Fehler'     , 10, 'center')
        self.add('Treffer'    , 10, 'center')
        self.add('Laufzeit'   , 10, 'center') 
        self.add('Strafen'    , 10, 'center')
        self.add('Einheit'    , 10, 'center')
        
        if team >= 2:
            for cnt in range(1,team+1):
                self.copy('Name'       , cnt, team)
                self.copy('Verein'     , cnt, team) 
                self.copy('Klasse'     , cnt, team) 
                self.copy('Zeit'       , cnt, team) 
                self.copy('Fehler'     , cnt, team) 
                self.copy('Treffer'    , cnt, team)
                self.copy('Laufzeit'   , cnt, team) 
                self.copy('Strafen'    , cnt, team)
                self.copy('Einheit'    , cnt, team)
    
    def add(self, name, width, halign):
        self.column[name] = Column(name, width, Alignment(vertical='center', horizontal=halign))
        
    def copy(self, key, cnt, team):
        self.n_staffel = max(self.n_staffel, team)
        column = self.column[key]
        name   = "%s %d/%d" % (key, cnt, team)
        self.column[name] = column.copy(name)
        
    def write(self, text, data, starter=None, newline=False):
        if newline: self.row += 1
        key = text if starter is None else "%s %d/%d" % (text, starter, self.n_staffel)
        self.column[key].write(self.row, data)

    def max_col(self):
        return max( c.num for c in self.column.values() )
        
        
def medaillenspiegel(ms, writer, toprule, stand, style):
    writer.cell(1,1, 'Medaillenspiegel', style['huge'])
    
    writer.cell(3,1, f"{ms.starter} Starter:innen bei {ms.meldungen} Meldungen aus {ms.vereine} Vereinen")
    
    toprule(5)
    writer.cell(5,2, "Verein", style['bold'])
    writer.cell(5,3, "Gold"  , style['bold'], style['center'])
    writer.cell(5,4, "Silber", style['bold'], style['center'])
    writer.cell(5,5, "Bronze", style['bold'], style['center'])
    toprule(6)

    width = 10    
    for row, verein in enumerate(ms.ergebnisse,6):
        width = max(width, len(verein.verein))
        if verein.first:
            writer.cell(row, 1, verein.position, style['center'])
        writer.cell(row, 2, verein.verein)
        writer.cell(row, 3, verein._gold  , style['center'])
        writer.cell(row, 4, verein._silber, style['center'])
        writer.cell(row, 5, verein._bronze, style['center'])
    
    stand(row)
    l = len(ms.info)
    idx = ms.info.find(' ', int(l/2))
    writer.cell(row+3, 1, ms.info[:idx], style['tiny'])
    writer.cell(row+4, 1, ms.info[idx:], style['tiny'])
    
    return width+3

def generic_export(wettkampf_oder_lauf, header, writer, 
           toprule = lambda row, start=1, stop=9: None, 
           stand   = lambda row, start=1, stop=9: None, 
           style   = defaultdict(int),
           empty   = True, tag     = None ):
    
    writer.cell(1,1, header, style['huge'])
    
    row = write_header(writer.cell, style['center'])

    for wertung in collect_data(wettkampf_oder_lauf, empty, tag):
        writer.staffel_mode = StaffelMode.Start if wertung.ist_staffel else StaffelMode.Off
        pos, sieger, row = 1, None, write_klasse(row, wertung.klasse, writer.cell)
        toprule(row)

        for team in Team.sortiere(wertung.teams):
            pos = write_platz(row, team, pos, writer.cell, style)
            write_name_verein(row, team, writer.cell)

            if team.platz:
                sieger = sieger or team.zeit()
                write_result(row, team, sieger, writer.cell, style)
                
            row += 1
            if team.ist_staffel():
                toprule(row-1, start=2)
                for s in team.liste():
                    writer.cell(row, 3, s.get_name(), style['tiny'], style['vcenter'])
                    writer.cell(row, 4, s.verein    , style['tiny'], style['vcenter'])
                    if team.has_finished():
                        zeit    = time2str(s.zeit())
                        fehler  = s.fehler or 0
                        writer.cell(row, 5, f"{zeit} [{fehler}]", style['tiny'], style['center' ])
                        if s.strafen > 0:
                            strafe = f"{s.strafen}x{s.einheit}s = {time2str(s.strafen*s.einheit,zehntel=False)}"
                            writer.cell(row, 9, strafe, style['tiny'], style['center' ])
                    else:
                        writer.cell(row, 5, "")
                    row += 1
  
    stand(row, stop=9)
    
def write_header(write_cell, center):
    row = 3
    write_cell(row,6,"Zeit"    , center)
    write_cell(row,7,"Abstand" , center)
    write_cell(row,8,"Fehler"  , center)
    return row

def write_name_verein(row, team, write_cell):
    name, verein = team.get_name_verein()
    write_cell(row, 3, name)
    write_cell(row, 4, verein)

def write_klasse(row, key, write_cell):
    write_cell(row+1,1, key)
    return row+1

def write_platz(row, team, pos, write_cell, style):
    if   team.is_dsq():
        write_cell(row, 2, "DSQ")
    elif team.platz:
        write_cell(row, 2, ("%d." % pos) if team.wertung else "--", style['center'])
    elif team.lauf.has_finished():
        write_cell(row, 2, "DNF")
    return pos+(1 if team.wertung and not team.is_dsq() else 0)

def write_result(row, team, sieger, write_cell, style):
    write_cell(row, 6, time2str(team.zeit()), style['center'])
    
    if sieger is not None:
        delta = team.zeit() - sieger
        if delta > 0:
            abstand = time2str(delta)
            write_cell(row, 7, abstand, style['center'])

    write_cell(row, 8, team.fehler() , style['center'])
    write_cell(row, 9, team.strafen(), style['tiny'], style['center'])
    